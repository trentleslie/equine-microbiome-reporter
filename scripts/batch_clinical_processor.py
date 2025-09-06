#!/usr/bin/env python3
"""
Batch Clinical Processing for Multiple FASTQ Files

Processes multiple samples in batch, generating individual and combined reports.
Optimized for HippoVet+ workflow of ~15 samples per week.

Usage:
    python batch_clinical_processor.py --input /path/to/epi2me/output --output /path/to/results
"""

import os
import sys
import pandas as pd
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import concurrent.futures
from tqdm import tqdm
import time

# Add src to path
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from clinical_filter import ClinicalFilter
from curation_interface import CurationInterface
from kraken2_classifier import Kraken2Classifier


class BatchClinicalProcessor:
    """Batch processor for multiple FASTQ files with clinical filtering."""
    
    def __init__(self, output_dir: Path = None, parallel: bool = True):
        """
        Initialize batch processor.
        
        Args:
            output_dir: Output directory for results
            parallel: Enable parallel processing
        """
        self.output_dir = output_dir or Path("/mnt/c/Users/hippovet/Desktop/batch_results")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.parallel = parallel
        self.filter = ClinicalFilter()
        self.curator = CurationInterface()
        
        # Track processing statistics
        self.stats = {
            'total_samples': 0,
            'successful': 0,
            'failed': 0,
            'total_species_before': 0,
            'total_species_after': 0,
            'processing_time': 0,
            'samples_processed': []
        }
    
    def find_all_samples(self, input_dir: Path) -> Dict[str, Dict[str, Path]]:
        """
        Find all samples and their associated files.
        
        Args:
            input_dir: Epi2Me output directory
            
        Returns:
            Dictionary mapping barcodes to file paths
        """
        samples = {}
        
        # Look for Kraken2 reports
        kraken_dir = input_dir / "kraken2-classified"
        if kraken_dir.exists():
            for report_file in kraken_dir.glob("*.kreport*"):
                barcode = report_file.stem.split('.')[0]
                if barcode not in samples:
                    samples[barcode] = {}
                samples[barcode]['kraken_report'] = report_file
        
        # Look for FASTQ files
        fastq_dir = input_dir / "fastq_pass"
        if fastq_dir.exists():
            for barcode_dir in fastq_dir.iterdir():
                if barcode_dir.is_dir():
                    barcode = barcode_dir.name
                    fastq_files = list(barcode_dir.glob("*.fastq*"))
                    if fastq_files:
                        if barcode not in samples:
                            samples[barcode] = {}
                        samples[barcode]['fastq_files'] = fastq_files
        
        print(f"Found {len(samples)} samples to process")
        for barcode in sorted(samples.keys()):
            files = samples[barcode]
            print(f"  {barcode}: {len(files.get('fastq_files', []))} FASTQ files")
        
        return samples
    
    def process_single_sample(self, barcode: str, files: Dict[str, any], 
                             database: str) -> Tuple[str, pd.DataFrame, Dict]:
        """
        Process a single sample with clinical filtering.
        
        Args:
            barcode: Sample barcode
            files: Dictionary of file paths
            database: Database name
            
        Returns:
            Tuple of (barcode, filtered_dataframe, statistics)
        """
        start_time = time.time()
        sample_stats = {
            'barcode': barcode,
            'database': database,
            'species_before': 0,
            'species_after': 0,
            'processing_time': 0,
            'status': 'pending'
        }
        
        try:
            # Parse Kraken2 report if available
            if 'kraken_report' in files:
                df = self._parse_kraken_report(files['kraken_report'])
                sample_stats['species_before'] = len(df)
                
                # Apply clinical filtering
                df_filtered = self.filter.process_results(
                    df, database, auto_exclude=True
                )
                sample_stats['species_after'] = len(df_filtered)
                
                # Add barcode column
                df_filtered['barcode'] = barcode
                
                sample_stats['status'] = 'success'
                sample_stats['processing_time'] = time.time() - start_time
                
                return barcode, df_filtered, sample_stats
            
            else:
                sample_stats['status'] = 'no_kraken_report'
                return barcode, pd.DataFrame(), sample_stats
                
        except Exception as e:
            sample_stats['status'] = f'error: {str(e)}'
            sample_stats['processing_time'] = time.time() - start_time
            return barcode, pd.DataFrame(), sample_stats
    
    def _parse_kraken_report(self, report_path: Path) -> pd.DataFrame:
        """Parse Kraken2 report file."""
        data = []
        with open(report_path, 'r') as f:
            for line in f:
                if line.strip():
                    parts = line.strip().split('\t')
                    if len(parts) >= 6:
                        percentage = float(parts[0])
                        reads_assigned = int(parts[2])
                        rank = parts[3]
                        name = parts[5].strip()
                        
                        # Only species-level
                        if rank == 'S':
                            data.append({
                                'species': name,
                                'percentage': percentage,
                                'abundance_reads': reads_assigned
                            })
        
        return pd.DataFrame(data)
    
    def process_batch(self, input_dir: Path, databases: List[str] = None,
                     max_workers: int = 4) -> Dict:
        """
        Process all samples in batch.
        
        Args:
            input_dir: Epi2Me output directory
            databases: List of databases to process
            max_workers: Maximum parallel workers
            
        Returns:
            Processing results dictionary
        """
        databases = databases or ['PlusPFP-16', 'EuPathDB', 'Viral']
        
        print(f"\n{'='*60}")
        print(f"Starting Batch Processing")
        print(f"Input: {input_dir}")
        print(f"Databases: {', '.join(databases)}")
        print(f"{'='*60}\n")
        
        # Find all samples
        samples = self.find_all_samples(input_dir)
        self.stats['total_samples'] = len(samples)
        
        batch_results = {}
        
        for database in databases:
            print(f"\n📊 Processing {database} database...")
            
            # Create database-specific output directory
            db_output = self.output_dir / database
            db_output.mkdir(exist_ok=True)
            
            all_results = []
            all_stats = []
            
            if self.parallel and max_workers > 1:
                # Parallel processing
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = {
                        executor.submit(
                            self.process_single_sample, 
                            barcode, files, database
                        ): barcode 
                        for barcode, files in samples.items()
                    }
                    
                    # Progress bar
                    with tqdm(total=len(futures), desc=f"Processing {database}") as pbar:
                        for future in concurrent.futures.as_completed(futures):
                            barcode, df_result, stats = future.result()
                            all_results.append(df_result)
                            all_stats.append(stats)
                            pbar.update(1)
                            
                            if stats['status'] == 'success':
                                self.stats['successful'] += 1
                            else:
                                self.stats['failed'] += 1
            else:
                # Sequential processing
                for barcode, files in tqdm(samples.items(), desc=f"Processing {database}"):
                    barcode, df_result, stats = self.process_single_sample(
                        barcode, files, database
                    )
                    all_results.append(df_result)
                    all_stats.append(stats)
                    
                    if stats['status'] == 'success':
                        self.stats['successful'] += 1
                    else:
                        self.stats['failed'] += 1
            
            # Combine results
            if all_results:
                df_combined = pd.concat(all_results, ignore_index=True)
                
                # Export individual Excel files for review
                for barcode in samples.keys():
                    df_barcode = df_combined[df_combined['barcode'] == barcode]
                    if not df_barcode.empty:
                        excel_path = self._export_sample_for_review(
                            df_barcode, barcode, database, db_output
                        )
                
                # Export combined results
                combined_path = db_output / f"combined_{database}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                df_combined.to_excel(combined_path, index=False)
                print(f"  ✅ Combined results: {combined_path.name}")
                
                batch_results[database] = {
                    'combined_df': df_combined,
                    'stats': all_stats,
                    'output_path': combined_path
                }
        
        # Generate batch report
        report_path = self._generate_batch_report(batch_results)
        
        print(f"\n{'='*60}")
        print(f"Batch Processing Complete!")
        print(f"Total samples: {self.stats['total_samples']}")
        print(f"Successful: {self.stats['successful']}")
        print(f"Failed: {self.stats['failed']}")
        print(f"Report: {report_path}")
        print(f"{'='*60}\n")
        
        return batch_results
    
    def _export_sample_for_review(self, df: pd.DataFrame, barcode: str, 
                                  database: str, output_dir: Path) -> Path:
        """Export individual sample for Excel review."""
        if df.empty:
            return None
            
        # Create Excel with clinical review format
        excel_path = output_dir / f"{barcode}_{database}_review.xlsx"
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Review sheet
            review_df = self.curator.create_review_summary(df, database)
            review_df.to_excel(writer, sheet_name='Review', index=False)
            
            # Format the review sheet
            worksheet = writer.sheets['Review']
            
            # Add conditional formatting for clinical relevance
            from openpyxl.styles import PatternFill
            
            for row in range(2, len(review_df) + 2):
                relevance_cell = worksheet.cell(row=row, column=4)
                if relevance_cell.value == 'high':
                    fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif relevance_cell.value == 'moderate':
                    fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
                elif relevance_cell.value == 'low':
                    fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                else:
                    fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                
                for col in range(1, 8):
                    worksheet.cell(row=row, column=col).fill = fill
        
        return excel_path
    
    def _generate_batch_report(self, batch_results: Dict) -> Path:
        """Generate comprehensive batch processing report."""
        report_path = self.output_dir / f"batch_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Batch Processing Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #2c3e50; }}
        h2 {{ color: #34495e; border-bottom: 2px solid #ecf0f1; padding-bottom: 5px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #3498db; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        .success {{ color: green; font-weight: bold; }}
        .failed {{ color: red; font-weight: bold; }}
        .summary-box {{ background-color: #ecf0f1; padding: 15px; border-radius: 5px; margin: 20px 0; }}
    </style>
</head>
<body>
    <h1>Batch Clinical Filtering Report</h1>
    <div class="summary-box">
        <h2>Summary</h2>
        <p><strong>Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>Total Samples:</strong> {self.stats['total_samples']}</p>
        <p><strong>Successful:</strong> <span class="success">{self.stats['successful']}</span></p>
        <p><strong>Failed:</strong> <span class="failed">{self.stats['failed']}</span></p>
    </div>
"""
        
        for database, results in batch_results.items():
            df = results['combined_df']
            stats = results['stats']
            
            html_content += f"""
    <h2>{database} Database</h2>
    <table>
        <tr>
            <th>Barcode</th>
            <th>Species Before</th>
            <th>Species After</th>
            <th>Reduction %</th>
            <th>Processing Time</th>
            <th>Status</th>
        </tr>
"""
            
            for stat in stats:
                if stat['species_before'] > 0:
                    reduction = (1 - stat['species_after']/stat['species_before']) * 100
                else:
                    reduction = 0
                
                status_class = 'success' if stat['status'] == 'success' else 'failed'
                
                html_content += f"""
        <tr>
            <td>{stat['barcode']}</td>
            <td>{stat['species_before']}</td>
            <td>{stat['species_after']}</td>
            <td>{reduction:.1f}%</td>
            <td>{stat['processing_time']:.2f}s</td>
            <td class="{status_class}">{stat['status']}</td>
        </tr>
"""
            
            html_content += """
    </table>
"""
            
            # Add species summary
            if not df.empty and 'clinical_relevance' in df.columns:
                relevance_counts = df['clinical_relevance'].value_counts()
                html_content += f"""
    <h3>Clinical Relevance Distribution</h3>
    <ul>
"""
                for relevance, count in relevance_counts.items():
                    html_content += f"        <li>{relevance.capitalize()}: {count} species</li>\n"
                html_content += "    </ul>\n"
        
        html_content += """
</body>
</html>
"""
        
        with open(report_path, 'w') as f:
            f.write(html_content)
        
        return report_path
    
    def create_combined_report(self, batch_results: Dict) -> Path:
        """
        Create a single combined Excel report for all databases and samples.
        
        Args:
            batch_results: Results from batch processing
            
        Returns:
            Path to combined report
        """
        report_path = self.output_dir / f"combined_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = []
            for database, results in batch_results.items():
                for stat in results['stats']:
                    summary_data.append({
                        'Database': database,
                        'Barcode': stat['barcode'],
                        'Species Before': stat['species_before'],
                        'Species After': stat['species_after'],
                        'Status': stat['status']
                    })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Individual database sheets
            for database, results in batch_results.items():
                df = results['combined_df']
                if not df.empty:
                    # Pivot by barcode for easier review
                    pivot_df = df.pivot_table(
                        values='abundance_reads',
                        index='species',
                        columns='barcode',
                        fill_value=0,
                        aggfunc='sum'
                    )
                    pivot_df.to_excel(writer, sheet_name=database[:30])  # Excel sheet name limit
        
        print(f"📊 Combined report created: {report_path}")
        return report_path


def main():
    """Main entry point for batch processor."""
    parser = argparse.ArgumentParser(
        description="Batch process multiple samples with clinical filtering"
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Epi2Me output directory containing all samples"
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output directory for results"
    )
    parser.add_argument(
        "--databases",
        nargs='+',
        choices=['PlusPFP-16', 'EuPathDB', 'Viral'],
        default=['PlusPFP-16', 'EuPathDB', 'Viral'],
        help="Databases to process (default: all)"
    )
    parser.add_argument(
        "--parallel",
        action="store_true",
        default=True,
        help="Enable parallel processing (default: True)"
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Number of parallel workers (default: 4)"
    )
    parser.add_argument(
        "--combined-report",
        action="store_true",
        help="Generate combined Excel report"
    )
    
    args = parser.parse_args()
    
    # Initialize processor
    processor = BatchClinicalProcessor(
        output_dir=args.output,
        parallel=args.parallel
    )
    
    # Process batch
    batch_results = processor.process_batch(
        input_dir=args.input,
        databases=args.databases,
        max_workers=args.workers
    )
    
    # Generate combined report if requested
    if args.combined_report:
        processor.create_combined_report(batch_results)
    
    # Print final statistics
    print("\n📈 Processing Statistics:")
    print(f"  Total samples: {processor.stats['total_samples']}")
    print(f"  Successful: {processor.stats['successful']}")
    print(f"  Failed: {processor.stats['failed']}")
    print(f"  Success rate: {processor.stats['successful']/max(processor.stats['total_samples']*len(args.databases), 1)*100:.1f}%")


if __name__ == "__main__":
    main()