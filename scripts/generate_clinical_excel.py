#!/usr/bin/env python3
"""
Enhanced Excel Review File Generator for Clinical Assessment
Creates color-coded Excel files with clinical relevance assessment
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ClinicalExcelGenerator:
    """Generate enhanced Excel review files with clinical assessment."""
    
    def __init__(self):
        """Initialize with clinical relevance categories."""
        self.clinical_categories = {
            'HIGH': {
                'color': 'FFE74C3C',  # Red
                'species': [
                    'Streptococcus equi', 'Rhodococcus equi', 'Salmonella',
                    'Clostridium difficile', 'Clostridium perfringens'
                ]
            },
            'MODERATE': {
                'color': 'FFF39C12',  # Orange
                'species': [
                    'Escherichia coli', 'Klebsiella', 'Enterococcus',
                    'Pseudomonas aeruginosa', 'Staphylococcus aureus'
                ]
            },
            'LOW': {
                'color': 'FF27AE60',  # Green
                'species': [
                    'Lactobacillus', 'Bifidobacterium', 'Faecalibacterium',
                    'Bacteroides fragilis', 'Prevotella'
                ]
            },
            'EXCLUDED': {
                'color': 'FF95A5A6',  # Gray
                'keywords': ['Phytophthora', 'plant', 'Arabidopsis', 'Oryza']
            }
        }
        
    def assess_clinical_relevance(self, species_name: str, abundance: float) -> Dict:
        """
        Assess clinical relevance of a species.
        
        Args:
            species_name: Species name
            abundance: Abundance percentage
            
        Returns:
            Dict with relevance category and recommendation
        """
        # Check for exclusions first
        for keyword in self.clinical_categories['EXCLUDED']['keywords']:
            if keyword.lower() in species_name.lower():
                return {
                    'category': 'EXCLUDED',
                    'include': 'NO',
                    'notes': 'Plant parasite - excluded',
                    'color': self.clinical_categories['EXCLUDED']['color']
                }
        
        # Check HIGH relevance pathogens
        for pathogen in self.clinical_categories['HIGH']['species']:
            if pathogen.lower() in species_name.lower():
                return {
                    'category': 'HIGH',
                    'include': 'YES',
                    'notes': f'Known pathogen (abundance: {abundance:.2f}%)',
                    'color': self.clinical_categories['HIGH']['color']
                }
        
        # Check MODERATE relevance
        for species in self.clinical_categories['MODERATE']['species']:
            if species.lower() in species_name.lower():
                include = 'YES' if abundance > 5.0 else '[Review]'
                return {
                    'category': 'MODERATE',
                    'include': include,
                    'notes': f'Opportunistic pathogen (abundance: {abundance:.2f}%)',
                    'color': self.clinical_categories['MODERATE']['color']
                }
        
        # Check beneficial bacteria
        for species in self.clinical_categories['LOW']['species']:
            if species.lower() in species_name.lower():
                return {
                    'category': 'LOW',
                    'include': 'YES' if abundance > 1.0 else 'NO',
                    'notes': f'Beneficial bacteria (abundance: {abundance:.2f}%)',
                    'color': self.clinical_categories['LOW']['color']
                }
        
        # Default assessment based on abundance
        if abundance > 10.0:
            return {
                'category': 'REVIEW',
                'include': '[Review]',
                'notes': f'High abundance - manual review needed',
                'color': 'FFFFFF00'  # Yellow
            }
        elif abundance > 1.0:
            return {
                'category': 'LOW',
                'include': 'NO',
                'notes': f'Low clinical relevance',
                'color': 'FFFFFFFF'  # White
            }
        else:
            return {
                'category': 'TRACE',
                'include': 'NO',
                'notes': f'Trace amount (<1%)',
                'color': 'FFF8F9F9'  # Light gray
            }
    
    def generate_review_excel(self, csv_file: Path, output_file: Path,
                             sample_name: str = "Unknown") -> Path:
        """
        Generate enhanced Excel review file from CSV data.
        
        Args:
            csv_file: Input CSV file with species data
            output_file: Output Excel file path
            sample_name: Sample identifier
            
        Returns:
            Path to generated Excel file
        """
        # Read CSV data
        df = pd.read_csv(csv_file)
        
        # Find barcode column
        barcode_cols = [col for col in df.columns if col.startswith('barcode')]
        if not barcode_cols:
            raise ValueError("No barcode column found in CSV")
        barcode_col = barcode_cols[0]
        
        # Calculate abundances
        total_reads = df[barcode_col].sum()
        df['abundance_pct'] = (df[barcode_col] / total_reads * 100).round(2)
        
        # Assess clinical relevance for each species
        assessments = []
        for _, row in df.iterrows():
            species = row.get('species', 'Unknown')
            abundance = row['abundance_pct']
            assessment = self.assess_clinical_relevance(species, abundance)
            
            assessments.append({
                'Species': species,
                'Abundance (%)': abundance,
                'Reads': row[barcode_col],
                'Phylum': row.get('phylum', 'Unknown'),
                'Genus': row.get('genus', 'Unknown'),
                'Clinical Relevance': assessment['category'],
                'Include in Report': assessment['include'],
                'Notes': assessment['notes'],
                '_color': assessment['color']
            })
        
        # Create review dataframe
        review_df = pd.DataFrame(assessments)
        review_df = review_df.sort_values('Abundance (%)', ascending=False)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        
        # Sheet 1: Clinical Review
        ws1 = wb.active
        ws1.title = "Clinical Review"
        
        # Add headers
        headers = ['Species', 'Abundance (%)', 'Reads', 'Phylum', 'Genus', 
                  'Clinical Relevance', 'Include in Report', 'Notes']
        ws1.append(headers)
        
        # Format headers
        header_fill = PatternFill(start_color='FF2C3E50', end_color='FF2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFFFF', bold=True)
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data with color coding
        for _, row in review_df.iterrows():
            row_data = [row[col] for col in headers]
            ws1.append(row_data)
            
            # Apply color to the entire row based on clinical relevance
            row_num = ws1.max_row
            fill = PatternFill(start_color=row['_color'], end_color=row['_color'], fill_type='solid')
            for cell in ws1[row_num]:
                if row['_color'] != 'FFFFFFFF':  # Don't fill white cells
                    cell.fill = fill
        
        # Adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Summary Statistics
        ws2 = wb.create_sheet("Summary")
        
        summary_data = [
            ['Metric', 'Value'],
            ['Sample Name', sample_name],
            ['Total Species', len(df)],
            ['Species After Filtering', len(review_df[review_df['Include in Report'] != 'NO'])],
            ['High Relevance Species', len(review_df[review_df['Clinical Relevance'] == 'HIGH'])],
            ['Moderate Relevance Species', len(review_df[review_df['Clinical Relevance'] == 'MODERATE'])],
            ['Excluded Species', len(review_df[review_df['Clinical Relevance'] == 'EXCLUDED'])],
            ['Manual Review Required', len(review_df[review_df['Include in Report'] == '[Review]'])],
            ['Reduction %', f"{(1 - len(review_df[review_df['Include in Report'] != 'NO']) / len(df)) * 100:.1f}%"]
        ]
        
        for row in summary_data:
            ws2.append(row)
        
        # Format summary headers
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Sheet 3: Instructions
        ws3 = wb.create_sheet("Instructions")
        
        instructions = [
            ['Clinical Review Instructions'],
            [''],
            ['Color Coding:'],
            ['🔴 RED - High clinical relevance (known pathogens)'],
            ['🟠 ORANGE - Moderate relevance (opportunistic pathogens)'],
            ['🟢 GREEN - Low relevance (beneficial bacteria)'],
            ['⚪ GRAY - Excluded (plant parasites, contaminants)'],
            ['🟡 YELLOW - Requires manual review'],
            [''],
            ['Review Process:'],
            ['1. Check all entries marked as [Review] in "Include in Report" column'],
            ['2. Modify "Include in Report" to YES or NO based on clinical judgment'],
            ['3. Add additional notes as needed'],
            ['4. Save the file when review is complete'],
            [''],
            ['Thresholds:'],
            ['- High abundance: >10%'],
            ['- Moderate abundance: 5-10%'],
            ['- Low abundance: 1-5%'],
            ['- Trace: <1%']
        ]
        
        for row in instructions:
            ws3.append(row if isinstance(row, list) else [row])
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Generated clinical review Excel: {output_file}")
        
        return output_file


def main():
    """Test the clinical Excel generator."""
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python generate_clinical_excel.py <csv_file> <output_excel>")
        sys.exit(1)
    
    csv_file = Path(sys.argv[1])
    output_file = Path(sys.argv[2])
    
    generator = ClinicalExcelGenerator()
    generator.generate_review_excel(csv_file, output_file, sample_name=csv_file.stem)
    print(f"Clinical review Excel generated: {output_file}")


if __name__ == "__main__":
    main()