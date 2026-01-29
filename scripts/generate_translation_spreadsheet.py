#!/usr/bin/env python3
"""
Generate a translation comparison spreadsheet for Gosia.

Translates all strings from config/translations.yaml using Google Translate,
DeepL, and (optionally) Claude API, then outputs an Excel file where Gosia can
review, compare, and finalize translations for Polish and German.

Requirements:
  - deep-translator (pip install deep-translator) — for Google Translate (free)
  - DEEPL_API_KEY environment variable (free tier: https://www.deepl.com/pro-api)
  - ANTHROPIC_API_KEY environment variable (optional, for Claude comparisons)

Usage:
  poetry run python scripts/generate_translation_spreadsheet.py
  poetry run python scripts/generate_translation_spreadsheet.py --output translations_for_review.xlsx
  poetry run python scripts/generate_translation_spreadsheet.py --no-claude --no-deepl  # Google only
  poetry run python scripts/generate_translation_spreadsheet.py --google-only  # Same as above
  poetry run python scripts/generate_translation_spreadsheet.py --deepl-only   # DeepL only
  poetry run python scripts/generate_translation_spreadsheet.py --claude-only   # Claude only
"""

import os
import sys
import argparse
import time
from pathlib import Path

import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

sys.path.append(str(Path(__file__).parent.parent))

load_dotenv()


def load_english_strings() -> dict:
    """Load English strings from translations.yaml."""
    config_path = Path(__file__).parent.parent / "config" / "translations.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        translations = yaml.safe_load(f)
    return translations.get('en', {})


def translate_with_google(texts: dict, target_lang: str) -> dict:
    """Translate texts using Google Translate via deep-translator (free, no API key).

    Args:
        texts: Dict of key -> English text
        target_lang: Target language code (pl, de)

    Returns:
        Dict of key -> translated text
    """
    try:
        from deep_translator import GoogleTranslator
    except ImportError:
        print("deep-translator not installed. Install with: pip install deep-translator")
        return {}

    results = {}

    for key, text in texts.items():
        try:
            translator = GoogleTranslator(source='en', target=target_lang.lower())
            result = translator.translate(text)
            results[key] = result
            time.sleep(0.2)  # Rate limiting
        except Exception as e:
            print(f"  Google error for '{key}': {e}")
            results[key] = f"[ERROR: {e}]"

    return results


def translate_with_deepl(texts: dict, target_lang: str, api_key: str) -> dict:
    """Translate texts using the official DeepL Python client.

    Args:
        texts: Dict of key -> English text
        target_lang: Target language code (PL, DE)
        api_key: DeepL API key

    Returns:
        Dict of key -> translated text
    """
    try:
        import deepl
    except ImportError:
        print("deepl not installed. Install with: pip install deepl")
        return {}

    results = {}
    translator = deepl.Translator(api_key)

    for key, text in texts.items():
        try:
            result = translator.translate_text(text, target_lang=target_lang.upper())
            results[key] = result.text
            time.sleep(0.1)  # Rate limiting
        except Exception as e:
            print(f"  DeepL error for '{key}': {e}")
            results[key] = f"[ERROR: {e}]"

    return results


def translate_with_claude(texts: dict, target_lang: str, api_key: str) -> dict:
    """Translate texts using Claude API with veterinary/scientific context.

    Args:
        texts: Dict of key -> English text
        target_lang: Target language name (Polish, German)
        api_key: Anthropic API key

    Returns:
        Dict of key -> translated text
    """
    try:
        import anthropic
    except ImportError:
        print("anthropic not installed. Install with: poetry install --with llm")
        return {}

    client = anthropic.Anthropic(api_key=api_key)
    results = {}

    # Batch texts into groups to reduce API calls
    batch_size = 20
    keys = list(texts.keys())

    for i in range(0, len(keys), batch_size):
        batch_keys = keys[i:i + batch_size]
        batch_texts = {k: texts[k] for k in batch_keys}

        # Build prompt
        lines = []
        for k, v in batch_texts.items():
            lines.append(f"{k}: {v}")
        text_block = "\n".join(lines)

        prompt = f"""Translate these veterinary microbiome report strings from English to {target_lang}.

Rules:
- Do NOT translate Latin scientific terms (species names, phylum names like Bacillota, Bacteroidota, Pseudomonadota, Actinomycetota)
- Do NOT translate technical terms: NGS, PCR, DNA, RNA, Illumina, Nanopore, NG-GP, 16S rRNA, dysbiosis, microbiome, microbiota
- Keep placeholders like {{current}}, {{total}} unchanged
- Keep email addresses and phone numbers unchanged
- Use formal/professional register appropriate for a medical/veterinary laboratory report

Strings to translate (format: key: text):
{text_block}

Return ONLY the translations in the same format (key: translated_text), one per line. No explanations."""

        try:
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}]
            )

            # Parse response
            response_text = message.content[0].text
            for line in response_text.strip().split('\n'):
                if ':' in line:
                    k, v = line.split(':', 1)
                    k = k.strip()
                    v = v.strip()
                    if k in batch_texts:
                        results[k] = v

            time.sleep(1)  # Rate limiting

        except Exception as e:
            print(f"  Claude error for batch {i//batch_size + 1}: {e}")
            for k in batch_keys:
                results[k] = f"[ERROR: {e}]"

    return results


def create_spreadsheet(
    english: dict,
    google_pl: dict,
    deepl_pl: dict,
    claude_pl: dict,
    google_de: dict,
    deepl_de: dict,
    claude_de: dict,
    output_path: str
):
    """Create formatted Excel spreadsheet for translation review.

    Columns: Key | English | Google_PL | DeepL_PL | Claude_PL | Final_PL |
             Google_DE | DeepL_DE | Claude_DE | Final_DE | Notes
    Final columns pre-populated with DeepL suggestion as default.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    final_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    key_font = Font(name="Consolas", size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Key", "English", "Google_PL", "DeepL_PL", "Claude_PL", "Final_PL",
               "Google_DE", "DeepL_DE", "Claude_DE", "Final_DE", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, (key, en_text) in enumerate(english.items(), 2):
        # For Final columns, prefer DeepL, fall back to Google
        final_pl = deepl_pl.get(key, '') or google_pl.get(key, '')
        final_de = deepl_de.get(key, '') or google_de.get(key, '')

        values = [
            key,
            en_text,
            google_pl.get(key, ''),
            deepl_pl.get(key, ''),
            claude_pl.get(key, ''),
            final_pl,
            google_de.get(key, ''),
            deepl_de.get(key, ''),
            claude_de.get(key, ''),
            final_de,
            '',  # Notes
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = wrap_alignment
            cell.border = thin_border

            if col == 1:
                cell.font = key_font
            if col in (6, 10):  # Final columns
                cell.fill = final_fill

    # Column widths
    widths = [30, 50, 45, 45, 45, 50, 45, 45, 45, 50, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "Translation Review Spreadsheet — Equine Microbiome Reporter",
        "",
        "Instructions:",
        "1. Review the 'Translations' sheet",
        "2. For each row, compare Google Translate, DeepL, and Claude suggestions",
        "3. Edit the Final_PL and Final_DE columns (green) with your preferred translation",
        "4. Use the Notes column to flag any issues or questions",
        "",
        "Translation Sources:",
        "- Google_PL / Google_DE: Google Translate (free, good baseline)",
        "- DeepL_PL / DeepL_DE: DeepL (high quality, especially for European languages)",
        "- Claude_PL / Claude_DE: Claude AI (context-aware, veterinary terminology)",
        "",
        "Rules:",
        "- Do NOT translate Latin scientific terms (Bacillota, Bacteroidota, etc.)",
        "- Do NOT translate: dysbiosis, microbiome, microbiota, NGS, PCR, 16S rRNA, NG-GP",
        "- Keep placeholders like {current}, {total} unchanged",
        "- Use formal/professional register for laboratory reports",
        "",
        "The Final columns are pre-populated with the DeepL suggestion (or Google if DeepL unavailable).",
        "You only need to edit what's wrong or could be improved.",
        "",
        "When done, save the file and run:",
        "  poetry run python scripts/import_translation_spreadsheet.py <this_file.xlsx>",
    ]
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
    ws2.column_dimensions['A'].width = 80

    wb.save(output_path)
    print(f"Spreadsheet saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate translation spreadsheet for review")
    parser.add_argument('--output', '-o', default='translations_for_review.xlsx',
                       help='Output Excel file path')
    # Exclusive source flags
    parser.add_argument('--google-only', action='store_true',
                       help='Use only Google Translate (skip DeepL and Claude)')
    parser.add_argument('--deepl-only', action='store_true',
                       help='Use only DeepL (skip Google and Claude)')
    parser.add_argument('--claude-only', action='store_true',
                       help='Use only Claude (skip Google and DeepL)')
    # Individual disable flags
    parser.add_argument('--no-google', action='store_true',
                       help='Skip Google Translate')
    parser.add_argument('--no-deepl', action='store_true',
                       help='Skip DeepL translations')
    parser.add_argument('--no-claude', action='store_true',
                       help='Skip Claude translations')
    args = parser.parse_args()

    # Resolve which services to use
    use_google = True
    use_deepl = True
    use_claude = True

    # *-only flags disable the other two
    if args.google_only:
        use_deepl = False
        use_claude = False
    elif args.deepl_only:
        use_google = False
        use_claude = False
    elif args.claude_only:
        use_google = False
        use_deepl = False

    # --no-* flags disable individually
    if args.no_google:
        use_google = False
    if args.no_deepl:
        use_deepl = False
    if args.no_claude:
        use_claude = False

    # Load English strings
    english = load_english_strings()
    print(f"Loaded {len(english)} English strings")

    # Get API keys
    deepl_key = os.environ.get('DEEPL_API_KEY', '')
    anthropic_key = os.environ.get('ANTHROPIC_API_KEY', '')

    google_pl = {}
    google_de = {}
    deepl_pl = {}
    deepl_de = {}
    claude_pl = {}
    claude_de = {}

    # Google Translate (free, no API key needed)
    if use_google:
        print("\nTranslating with Google Translate...")
        print("  Polish...")
        google_pl = translate_with_google(english, 'pl')
        print(f"  Got {len(google_pl)} Polish translations")

        print("  German...")
        google_de = translate_with_google(english, 'de')
        print(f"  Got {len(google_de)} German translations")
    else:
        print("\nSkipping Google Translate.")

    # DeepL translations
    if use_deepl:
        if deepl_key:
            print("\nTranslating with DeepL...")
            print("  Polish...")
            deepl_pl = translate_with_deepl(english, 'PL', deepl_key)
            print(f"  Got {len(deepl_pl)} Polish translations")

            print("  German...")
            deepl_de = translate_with_deepl(english, 'DE', deepl_key)
            print(f"  Got {len(deepl_de)} German translations")
        else:
            print("\nNo DEEPL_API_KEY found. Skipping DeepL translations.")
            print("Get a free API key at: https://www.deepl.com/pro-api")
    else:
        print("\nSkipping DeepL translations.")

    # Claude translations
    if use_claude:
        if anthropic_key:
            print("\nTranslating with Claude...")
            print("  Polish...")
            claude_pl = translate_with_claude(english, 'Polish', anthropic_key)
            print(f"  Got {len(claude_pl)} Polish translations")

            print("  German...")
            claude_de = translate_with_claude(english, 'German', anthropic_key)
            print(f"  Got {len(claude_de)} German translations")
        else:
            print("\nNo ANTHROPIC_API_KEY found. Skipping Claude translations.")
            print("Install with: poetry install --with llm")
    else:
        print("\nSkipping Claude translations.")

    # Check if we got any translations at all
    all_dicts = [google_pl, deepl_pl, claude_pl, google_de, deepl_de, claude_de]
    if not any(all_dicts):
        print("\nNo translations were generated. Creating spreadsheet with English-only columns.")
        print("Set DEEPL_API_KEY and/or ANTHROPIC_API_KEY, or use Google Translate (free).")

    # Create spreadsheet
    print(f"\nCreating spreadsheet: {args.output}")
    create_spreadsheet(english, google_pl, deepl_pl, claude_pl,
                      google_de, deepl_de, claude_de, args.output)

    # Summary
    counts = {
        'Google PL': len(google_pl), 'DeepL PL': len(deepl_pl), 'Claude PL': len(claude_pl),
        'Google DE': len(google_de), 'DeepL DE': len(deepl_de), 'Claude DE': len(claude_de),
    }
    total = sum(counts.values())
    print(f"\nDone! {total} translations across all services/languages.")
    for name, count in counts.items():
        if count > 0:
            print(f"  {name}: {count}")
    print(f"\nSend '{args.output}' to Gosia for review.")


if __name__ == '__main__':
    main()
