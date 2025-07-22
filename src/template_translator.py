"""
Template translation workflow manager.
Handles batch translation of Jinja2 templates with validation and review export.
"""

import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .translation_service import TranslationService, get_translation_service


class TemplateTranslationWorkflow:
    """Orchestrates the translation of all templates"""
    
    def __init__(self, 
                 project_root: Path,
                 translation_service: Optional[TranslationService] = None,
                 source_language: str = "en",
                 target_languages: List[str] = None):
        self.project_root = project_root
        self.templates_dir = project_root / "templates"
        self.translation_service = translation_service or get_translation_service("free")
        self.source_language = source_language
        self.target_languages = target_languages or ["pl", "ja"]
        
        # Language names for display
        self.language_names = {
            "en": "English",
            "pl": "Polish", 
            "ja": "Japanese"
        }
    
    def translate_all_templates(self, verbose: bool = True) -> Dict[str, List[str]]:
        """
        Translate all English templates to target languages.
        Returns dict of {language: [translated_files]}
        """
        source_dir = self.templates_dir / self.source_language
        results = {}
        
        for target_lang in self.target_languages:
            if verbose:
                print(f"\nTranslating templates to {self.language_names[target_lang]}...")
            
            target_dir = self.templates_dir / target_lang
            translated_files = []
            
            # Process all .j2 files
            for template_file in source_dir.rglob("*.j2"):
                relative_path = template_file.relative_to(source_dir)
                target_file = target_dir / relative_path
                
                # Create target directory if it doesn't exist
                target_file.parent.mkdir(parents=True, exist_ok=True)
                
                if verbose:
                    print(f"  Translating: {relative_path}")
                
                try:
                    # Translate the template
                    translated_content = self.translation_service.translate_template(
                        template_file, 
                        target_lang
                    )
                    
                    # Save translated template
                    with open(target_file, 'w', encoding='utf-8') as f:
                        f.write(translated_content)
                    
                    translated_files.append(str(relative_path))
                
                except Exception as e:
                    print(f"    Error: {e}")
            
            results[target_lang] = translated_files
        
        return results
    
    def validate_translations(self, verbose: bool = True) -> Dict[str, Dict[str, str]]:
        """
        Validate that translations preserve Jinja2 syntax.
        Returns dict of {language: {file: status}}
        """
        if verbose:
            print("\nValidating translations...")
        
        source_dir = self.templates_dir / self.source_language
        validation_results = {}
        
        for target_lang in self.target_languages:
            if verbose:
                print(f"\nValidating {self.language_names[target_lang]} translations:")
            
            target_dir = self.templates_dir / target_lang
            file_results = {}
            
            for source_file in source_dir.rglob("*.j2"):
                relative_path = source_file.relative_to(source_dir)
                target_file = target_dir / relative_path
                
                if not target_file.exists():
                    status = "Missing"
                    if verbose:
                        print(f"  ❌ Missing: {relative_path}")
                else:
                    # Extract Jinja2 elements from both files
                    try:
                        with open(source_file, 'r', encoding='utf-8') as f:
                            source_content = f.read()
                            source_jinja = re.findall(r'{{.*?}}|{%.*?%}|{#.*?#}', source_content)
                        
                        with open(target_file, 'r', encoding='utf-8') as f:
                            target_content = f.read()
                            target_jinja = re.findall(r'{{.*?}}|{%.*?%}|{#.*?#}', target_content)
                        
                        # Compare Jinja2 elements
                        if len(source_jinja) != len(target_jinja):
                            status = f"Mismatch (src: {len(source_jinja)}, tgt: {len(target_jinja)})"
                            if verbose:
                                print(f"  ⚠️  {relative_path}: Jinja2 element count mismatch")
                        else:
                            status = "Valid"
                            if verbose:
                                print(f"  ✅ {relative_path}: Valid")
                    
                    except Exception as e:
                        status = f"Error: {str(e)}"
                        if verbose:
                            print(f"  ❌ {relative_path}: Error reading file")
                
                file_results[str(relative_path)] = status
            
            validation_results[target_lang] = file_results
        
        return validation_results
    
    def create_review_spreadsheet(self, language: str, output_dir: Optional[Path] = None) -> Path:
        """Create an Excel spreadsheet for manual review of translations"""
        source_dir = self.templates_dir / self.source_language
        target_dir = self.templates_dir / language
        output_dir = output_dir or self.project_root
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Translations"
        
        # Add headers with styling
        headers = ['File', 'Line', 'English', self.language_names[language], 
                  'Approved', 'Corrected', 'Notes']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Collect translation data
        row_num = 2
        
        for source_file in source_dir.rglob("*.j2"):
            relative_path = source_file.relative_to(source_dir)
            target_file = target_dir / relative_path
            
            if target_file.exists():
                try:
                    with open(source_file, 'r', encoding='utf-8') as f:
                        source_lines = f.readlines()
                    
                    with open(target_file, 'r', encoding='utf-8') as f:
                        target_lines = f.readlines()
                    
                    # Match lines for review
                    for i, (src_line, tgt_line) in enumerate(zip(source_lines, target_lines)):
                        src_line = src_line.strip()
                        tgt_line = tgt_line.strip()
                        
                        # Only include translatable content (skip empty lines and pure Jinja2)
                        if src_line and not re.match(r'^({%|{{|{#).*?(}%|}}|#})$', src_line):
                            ws.cell(row=row_num, column=1, value=str(relative_path))
                            ws.cell(row=row_num, column=2, value=i + 1)
                            ws.cell(row=row_num, column=3, value=src_line)
                            ws.cell(row=row_num, column=4, value=tgt_line)
                            
                            # Add checkboxes for approval column
                            ws.cell(row=row_num, column=5, value="☐")
                            
                            # Alternate row coloring for readability
                            if row_num % 2 == 0:
                                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", 
                                                 fill_type="solid")
                                for col in range(1, 8):
                                    ws.cell(row=row_num, column=col).fill = fill
                            
                            row_num += 1
                
                except Exception as e:
                    print(f"Error processing {relative_path}: {e}")
        
        # Auto-adjust column widths
        column_widths = {
            'A': 30,  # File
            'B': 8,   # Line
            'C': 80,  # English
            'D': 80,  # Target language
            'E': 10,  # Approved
            'F': 80,  # Corrected
            'G': 40   # Notes
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        output_file = output_dir / f"translation_review_{language}.xlsx"
        wb.save(output_file)
        
        print(f"Created review spreadsheet: {output_file}")
        return output_file
    
    def apply_reviewed_translations(self, review_file: Path, language: str) -> int:
        """
        Apply corrections from reviewed spreadsheet back to templates.
        Returns number of corrections applied.
        """
        # Load reviewed spreadsheet
        df = pd.read_excel(review_file)
        corrections_applied = 0
        
        # Group by file
        for file_path, group in df.groupby('File'):
            target_file = self.templates_dir / language / file_path
            
            if target_file.exists():
                # Read current content
                with open(target_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                # Apply corrections
                for _, row in group.iterrows():
                    if pd.notna(row['Corrected']) and row['Corrected'].strip():
                        line_num = int(row['Line']) - 1
                        if 0 <= line_num < len(lines):
                            # Preserve leading/trailing whitespace
                            original_line = lines[line_num]
                            indent = len(original_line) - len(original_line.lstrip())
                            lines[line_num] = ' ' * indent + row['Corrected'].strip() + '\n'
                            corrections_applied += 1
                
                # Write back
                with open(target_file, 'w', encoding='utf-8') as f:
                    f.writelines(lines)
        
        print(f"Applied {corrections_applied} corrections from review")
        return corrections_applied
    
    def generate_translation_report(self, output_file: Optional[Path] = None) -> Path:
        """Generate a summary report of translation status"""
        output_file = output_file or self.project_root / "translation_report.md"
        
        # Validate all translations
        validation_results = self.validate_translations(verbose=False)
        
        # Generate report
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("# Translation Status Report\n\n")
            f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            for lang in self.target_languages:
                f.write(f"## {self.language_names[lang]} ({lang})\n\n")
                
                if lang in validation_results:
                    results = validation_results[lang]
                    total_files = len(results)
                    valid_files = sum(1 for status in results.values() if status == "Valid")
                    missing_files = sum(1 for status in results.values() if status == "Missing")
                    error_files = total_files - valid_files - missing_files
                    
                    f.write(f"- Total templates: {total_files}\n")
                    f.write(f"- Valid: {valid_files} ✅\n")
                    f.write(f"- Missing: {missing_files} ❌\n")
                    f.write(f"- Errors: {error_files} ⚠️\n\n")
                    
                    if missing_files > 0 or error_files > 0:
                        f.write("### Issues:\n\n")
                        for file, status in sorted(results.items()):
                            if status != "Valid":
                                f.write(f"- `{file}`: {status}\n")
                        f.write("\n")
        
        print(f"Generated translation report: {output_file}")
        return output_file