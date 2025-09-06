"""
Semi-Automated Curation Interface for Clinical Review

Provides a streamlined interface for reviewing and curating Kraken2 results
before final report generation. Designed for HippoVet+ workflow.
"""

import pandas as pd
import json
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class CurationInterface:
    """
    Interactive curation interface for reviewing filtered results.
    Supports batch review, annotation, and decision logging.
    """
    
    def __init__(self, output_dir: Path = None):
        """
        Initialize curation interface.
        
        Args:
            output_dir: Directory for saving curation decisions
        """
        self.output_dir = output_dir or Path("curation_results")
        self.output_dir.mkdir(exist_ok=True)
        self.decisions_log = []
        self.current_session = {
            'start_time': datetime.now().isoformat(),
            'decisions': [],
            'statistics': {}
        }
    
    def create_review_summary(self, df: pd.DataFrame, database: str) -> pd.DataFrame:
        """
        Create a simplified review summary for manual curation.
        
        Args:
            df: Filtered DataFrame with clinical annotations
            database: Database name
            
        Returns:
            Summary DataFrame optimized for review
        """
        # Select key columns for review
        review_columns = [
            'species', 'percentage', 'abundance_reads', 
            'clinical_relevance', 'curation_notes'
        ]
        
        available_columns = [col for col in review_columns if col in df.columns]
        df_review = df[available_columns].copy()
        
        # Add review status column
        df_review['review_status'] = 'pending'
        df_review['include_in_report'] = None
        df_review['reviewer_notes'] = ''
        
        # Group by clinical relevance for easier review
        df_review = df_review.sort_values(
            by=['clinical_relevance', 'percentage'],
            ascending=[True, False]
        )
        
        return df_review
    
    def export_for_excel_review(self, df: pd.DataFrame, database: str) -> Path:
        """
        Export results to Excel for manual review.
        
        Args:
            df: DataFrame with results
            database: Database name
            
        Returns:
            Path to exported Excel file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{database}_review_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Main review sheet
            df_review = self.create_review_summary(df, database)
            df_review.to_excel(writer, sheet_name='Review', index=False)
            
            # Add formatting
            worksheet = writer.sheets['Review']
            
            # Set column widths
            column_widths = {
                'A': 40,  # species
                'B': 12,  # percentage
                'C': 15,  # abundance_reads
                'D': 20,  # clinical_relevance
                'E': 50,  # curation_notes
                'F': 15,  # review_status
                'G': 20,  # include_in_report
                'H': 50   # reviewer_notes
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Color code by clinical relevance
            from openpyxl.styles import PatternFill
            
            relevance_colors = {
                'high': 'FFE6E6',      # Light red
                'moderate': 'FFF4E6',   # Light orange
                'low': 'E6F3FF',       # Light blue
                'excluded': 'F0F0F0'    # Light gray
            }
            
            for row in range(2, len(df_review) + 2):
                relevance_cell = worksheet.cell(row=row, column=4)
                relevance = relevance_cell.value
                
                if relevance in relevance_colors:
                    fill = PatternFill(start_color=relevance_colors[relevance],
                                     end_color=relevance_colors[relevance],
                                     fill_type='solid')
                    for col in range(1, 9):
                        worksheet.cell(row=row, column=col).fill = fill
            
            # Add instructions sheet
            instructions = pd.DataFrame({
                'Instructions': [
                    'Review Process:',
                    '1. Review each species in the "Review" sheet',
                    '2. Set "include_in_report" to YES or NO',
                    '3. Add any notes in "reviewer_notes" column',
                    '4. Save the file when complete',
                    '',
                    'Clinical Relevance Levels:',
                    '- HIGH: Known pathogens (usually include)',
                    '- MODERATE: Requires review based on abundance/context',
                    '- LOW: Usually commensal (include if abundant)',
                    '- EXCLUDED: Automatically filtered (plant parasites, etc.)',
                    '',
                    f'Database: {database}',
                    f'Total species: {len(df_review)}',
                    f'Generated: {timestamp}'
                ]
            })
            instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        logger.info(f"Exported review file: {filepath}")
        return filepath
    
    def import_excel_decisions(self, filepath: Path) -> pd.DataFrame:
        """
        Import curation decisions from reviewed Excel file.
        
        Args:
            filepath: Path to reviewed Excel file
            
        Returns:
            DataFrame with curation decisions
        """
        df_reviewed = pd.read_excel(filepath, sheet_name='Review')
        
        # Validate required columns
        required_columns = ['species', 'include_in_report']
        missing_columns = [col for col in required_columns if col not in df_reviewed.columns]
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        # Process decisions
        decisions = []
        for _, row in df_reviewed.iterrows():
            if pd.notna(row.get('include_in_report')):
                decision = {
                    'species': row['species'],
                    'include': str(row['include_in_report']).upper() == 'YES',
                    'percentage': row.get('percentage', 0),
                    'clinical_relevance': row.get('clinical_relevance', ''),
                    'reviewer_notes': row.get('reviewer_notes', ''),
                    'timestamp': datetime.now().isoformat()
                }
                decisions.append(decision)
        
        # Log decisions
        self.current_session['decisions'].extend(decisions)
        
        # Filter DataFrame based on decisions
        include_species = [d['species'] for d in decisions if d['include']]
        df_filtered = df_reviewed[df_reviewed['species'].isin(include_species)].copy()
        
        logger.info(f"Imported {len(decisions)} curation decisions")
        logger.info(f"Including {len(include_species)} species in report")
        
        return df_filtered
    
    def generate_batch_review_script(self, databases: List[str]) -> Path:
        """
        Generate a batch processing script for multiple databases.
        
        Args:
            databases: List of database names
            
        Returns:
            Path to batch script
        """
        script_content = f"""#!/usr/bin/env python3
\"\"\"
Batch Review Script for Kraken2 Results
Generated: {datetime.now().isoformat()}
\"\"\"

import pandas as pd
from pathlib import Path
import sys
sys.path.append('src')

from clinical_filter import ClinicalFilter
from curation_interface import CurationInterface

# Initialize components
filter_engine = ClinicalFilter()
curator = CurationInterface(output_dir=Path("curation_results"))

databases = {databases}

for database in databases:
    print(f"\\nProcessing {{database}}...")
    
    # Load raw results (update path as needed)
    input_file = f"results/{{database}}_kraken2.csv"
    if not Path(input_file).exists():
        print(f"  Skipping - file not found: {{input_file}}")
        continue
    
    df_raw = pd.read_csv(input_file)
    
    # Apply clinical filtering
    df_filtered = filter_engine.process_results(df_raw, database)
    
    # Export for review
    excel_path = curator.export_for_excel_review(df_filtered, database)
    print(f"  Exported for review: {{excel_path}}")
    
    # Generate curation report
    report = filter_engine.generate_curation_report(df_filtered, database)
    print(f"  Total species: {{report['total_species']}}")
    print(f"  Requires review: {{len(report['requires_review'])}}")
    print(f"  Auto-excluded: {{len(report['auto_excluded'])}}")

print("\\nBatch processing complete!")
print("Review Excel files in: curation_results/")
"""
        
        script_path = self.output_dir / "batch_review.py"
        with open(script_path, 'w') as f:
            f.write(script_content)
        
        script_path.chmod(0o755)  # Make executable
        logger.info(f"Generated batch review script: {script_path}")
        
        return script_path
    
    def create_decision_history(self, species: str) -> Dict:
        """
        Get historical curation decisions for a species.
        
        Args:
            species: Species name
            
        Returns:
            Dictionary with decision history
        """
        history_file = self.output_dir / "decision_history.json"
        
        if history_file.exists():
            with open(history_file, 'r') as f:
                all_history = json.load(f)
        else:
            all_history = {}
        
        species_history = all_history.get(species, {
            'total_reviews': 0,
            'included_count': 0,
            'excluded_count': 0,
            'last_decision': None,
            'notes': []
        })
        
        return species_history
    
    def update_decision_history(self, species: str, decision: bool, notes: str = "") -> None:
        """
        Update the decision history for a species.
        
        Args:
            species: Species name
            decision: Include (True) or exclude (False)
            notes: Optional reviewer notes
        """
        history_file = self.output_dir / "decision_history.json"
        
        if history_file.exists():
            with open(history_file, 'r') as f:
                all_history = json.load(f)
        else:
            all_history = {}
        
        if species not in all_history:
            all_history[species] = {
                'total_reviews': 0,
                'included_count': 0,
                'excluded_count': 0,
                'last_decision': None,
                'notes': []
            }
        
        # Update statistics
        all_history[species]['total_reviews'] += 1
        if decision:
            all_history[species]['included_count'] += 1
        else:
            all_history[species]['excluded_count'] += 1
        
        all_history[species]['last_decision'] = {
            'included': decision,
            'timestamp': datetime.now().isoformat(),
            'notes': notes
        }
        
        if notes:
            all_history[species]['notes'].append({
                'timestamp': datetime.now().isoformat(),
                'content': notes
            })
        
        # Save updated history
        with open(history_file, 'w') as f:
            json.dump(all_history, f, indent=2)
    
    def generate_curation_statistics(self) -> Dict:
        """
        Generate statistics from all curation sessions.
        
        Returns:
            Dictionary with curation statistics
        """
        history_file = self.output_dir / "decision_history.json"
        
        if not history_file.exists():
            return {'message': 'No curation history available'}
        
        with open(history_file, 'r') as f:
            all_history = json.load(f)
        
        stats = {
            'total_species_reviewed': len(all_history),
            'total_reviews': sum(h['total_reviews'] for h in all_history.values()),
            'consistency_score': 0,
            'most_included': [],
            'most_excluded': [],
            'requires_attention': []
        }
        
        # Calculate consistency score
        consistent_species = 0
        for species, history in all_history.items():
            if history['total_reviews'] > 0:
                include_rate = history['included_count'] / history['total_reviews']
                if include_rate == 1.0 or include_rate == 0.0:
                    consistent_species += 1
                elif 0.3 < include_rate < 0.7:
                    stats['requires_attention'].append(species)
        
        if all_history:
            stats['consistency_score'] = consistent_species / len(all_history)
        
        # Find most frequently included/excluded
        included_sorted = sorted(
            all_history.items(), 
            key=lambda x: x[1]['included_count'], 
            reverse=True
        )
        stats['most_included'] = [s[0] for s in included_sorted[:10] if s[1]['included_count'] > 0]
        
        excluded_sorted = sorted(
            all_history.items(), 
            key=lambda x: x[1]['excluded_count'], 
            reverse=True
        )
        stats['most_excluded'] = [s[0] for s in excluded_sorted[:10] if s[1]['excluded_count'] > 0]
        
        return stats